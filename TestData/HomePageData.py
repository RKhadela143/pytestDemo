import openpyxl


class HomePageData:
    test_homepage_data = [{"firstname":"Rahul", "lastname":"shetty", "gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:/Users/Radhe_Krishna/Desktop/Book1.xlsx")
        sheet = book.active
        dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dict]